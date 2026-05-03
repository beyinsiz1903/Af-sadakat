"""Report export router — PDF (reportlab) + Excel (openpyxl).

Endpoints:
  GET /api/v2/exports/tenants/{slug}/department-performance.{xlsx|pdf}
  GET /api/v2/exports/tenants/{slug}/sla-stats.{xlsx|pdf}
  GET /api/v2/exports/tenants/{slug}/loyalty-cohort.{xlsx|pdf}

Uses the tenant_guard dependency for auth and forwards to the existing report
fetcher functions so the data shown in the UI matches the export.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Iterable

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from core.tenant_guard import get_current_user

logger = logging.getLogger("omnihub.exports")
router = APIRouter(prefix="/api/v2/exports", tags=["exports"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _xlsx_response(filename: str, headers: list[str], rows: Iterable[list[Any]]) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2563EB")
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([("" if v is None else v) for v in r])
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(filename: str, title: str, headers: list[str], rows: Iterable[list[Any]]) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Paragraph(datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"), styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [headers] + [[("" if v is None else str(v)) for v in r] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _flatten_dept(payload: Any) -> tuple[list[str], list[list[Any]]]:
    headers = ["Department", "Total", "Resolved", "Avg Resolution (min)", "SLA Met %"]
    rows = []
    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        items = payload.get("departments") or payload.get("items") or []
    else:
        items = []
    for d in items:
        if not isinstance(d, dict):
            continue
        rows.append([
            d.get("department") or d.get("name") or "",
            d.get("total", d.get("count", 0)),
            d.get("resolved", 0),
            round(d.get("avg_resolution_minutes", d.get("avg_minutes", 0)) or 0, 1),
            round(d.get("sla_met_percent", d.get("sla_pct", 0)) or 0, 1),
        ])
    return headers, rows


def _flatten_sla(payload: dict) -> tuple[list[str], list[list[Any]]]:
    headers = ["Metric", "Value"]
    rows = []
    for k, v in payload.items():
        if isinstance(v, (int, float, str)):
            rows.append([k, v])
    if not rows:
        rows = [["(no data)", ""]]
    return headers, rows


def _flatten_cohort(payload: dict) -> tuple[list[str], list[list[Any]]]:
    headers = ["Cohort", "Members", "Active", "Retention %", "Revenue"]
    rows = []
    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        items = payload.get("cohorts") or payload.get("items") or []
    else:
        items = []
    for c in items:
        if not isinstance(c, dict):
            continue
        rows.append([
            c.get("cohort") or c.get("month") or "",
            c.get("members", c.get("size", 0)),
            c.get("active", 0),
            round(c.get("retention_percent", c.get("retention_pct", 0)) or 0, 1),
            c.get("revenue", 0),
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Department performance
# ---------------------------------------------------------------------------
async def _fetch_dept(slug: str) -> Any:
    try:
        from routers.reports import _build_department_performance  # type: ignore
        return await _build_department_performance(slug, 30)
    except Exception as e:
        logger.warning("dept fetch failed: %s", e)
        return []


@router.get("/tenants/{tenant_slug}/department-performance.xlsx")
async def export_dept_xlsx(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_dept(tenant_slug)
    headers, rows = _flatten_dept(payload)
    return _xlsx_response(f"dept_perf_{tenant_slug}.xlsx", headers, rows)


@router.get("/tenants/{tenant_slug}/department-performance.pdf")
async def export_dept_pdf(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_dept(tenant_slug)
    headers, rows = _flatten_dept(payload)
    return _pdf_response(f"dept_perf_{tenant_slug}.pdf", "Department Performance", headers, rows)


# ---------------------------------------------------------------------------
# SLA stats
# ---------------------------------------------------------------------------
async def _fetch_sla(slug: str) -> dict:
    try:
        from routers.sla import get_sla_stats  # type: ignore
        return await get_sla_stats(slug)
    except Exception:
        return {}


@router.get("/tenants/{tenant_slug}/sla-stats.xlsx")
async def export_sla_xlsx(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_sla(tenant_slug)
    headers, rows = _flatten_sla(payload)
    return _xlsx_response(f"sla_stats_{tenant_slug}.xlsx", headers, rows)


@router.get("/tenants/{tenant_slug}/sla-stats.pdf")
async def export_sla_pdf(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_sla(tenant_slug)
    headers, rows = _flatten_sla(payload)
    return _pdf_response(f"sla_stats_{tenant_slug}.pdf", "SLA Statistics", headers, rows)


# ---------------------------------------------------------------------------
# Loyalty cohort
# ---------------------------------------------------------------------------
async def _fetch_cohort(slug: str) -> dict:
    try:
        from routers.loyalty_analytics import cohort_analysis  # type: ignore
        return await cohort_analysis(slug)
    except Exception:
        return {"cohorts": []}


@router.get("/tenants/{tenant_slug}/loyalty-cohort.xlsx")
async def export_cohort_xlsx(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_cohort(tenant_slug)
    headers, rows = _flatten_cohort(payload)
    return _xlsx_response(f"loyalty_cohort_{tenant_slug}.xlsx", headers, rows)


@router.get("/tenants/{tenant_slug}/loyalty-cohort.pdf")
async def export_cohort_pdf(tenant_slug: str, user=Depends(get_current_user)):
    payload = await _fetch_cohort(tenant_slug)
    headers, rows = _flatten_cohort(payload)
    return _pdf_response(f"loyalty_cohort_{tenant_slug}.pdf", "Loyalty Cohort", headers, rows)
